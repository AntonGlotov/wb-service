from pandas import ExcelFile
from os import path
from sqlite3 import connect
from openpyxl import load_workbook


def export_to_sqlite(path_xls):
    clear_art_supp()

    base_name = 'database.sqlite3'
    con = connect(base_name)
    cursor = con.cursor()

    file_to_read = load_workbook(path_xls, data_only=True)

    xls = ExcelFile(path_xls)  # Получаем excel лист
    sheet_name = xls.sheet_names[0]
    sheet = file_to_read[sheet_name]

    for row in range(2, sheet.max_row + 1):
        data = []
        for col in range(1, 3):
            value = sheet.cell(row, col).value
            data.append(value)

        cursor.execute("INSERT INTO art_supp VALUES (?,?);", (data[0], data[1]))

    con.commit()
    con.close()


def clear_art_supp():

    prj_dir = path.abspath(path.curdir)

    base_name = 'database.sqlite3'
    con = connect(prj_dir + '/' + base_name)
    cursor = con.cursor()

    cursor.execute("DELETE FROM art_supp")
    con.commit()
    con.close()


if __name__ == '__main__':
    clear_art_supp()
    # export_to_sqlite('Articules.xlsx')

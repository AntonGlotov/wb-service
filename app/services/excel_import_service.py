import logging

from openpyxl import load_workbook
from pandas import ExcelFile

from app.repositories import article_supply_repository


logger = logging.getLogger(__name__)


def export_to_sqlite(path_xls):
    logger.info("Exporting article supplies from Excel to SQLite: path=%s", path_xls)
    rows = read_article_supply_rows(path_xls)
    article_supply_repository.replace(rows)
    logger.info("Article supplies exported to SQLite: path=%s rows_count=%s", path_xls, len(rows))


def read_article_supply_rows(path_xls):
    logger.info("Reading article supplies Excel file: path=%s", path_xls)
    try:
        file_to_read = load_workbook(path_xls, data_only=True)
        xls = ExcelFile(path_xls)
    except Exception:
        logger.exception("Failed to open article supplies Excel file: path=%s", path_xls)
        raise

    sheet_name = xls.sheet_names[0]
    sheet = file_to_read[sheet_name]
    logger.info("Reading article supplies sheet: path=%s sheet=%r max_row=%s", path_xls, sheet_name, sheet.max_row)

    rows = []
    empty_article_count = 0
    empty_supply_name_count = 0
    for row in range(2, sheet.max_row + 1):
        article = sheet.cell(row, 1).value
        supply_name = sheet.cell(row, 2).value
        if article is None:
            empty_article_count += 1
        if supply_name is None:
            empty_supply_name_count += 1
        if article is None or supply_name is None:
            logger.warning(
                "Article supplies row has empty values: path=%s sheet=%r row=%s article_empty=%s supply_name_empty=%s",
                path_xls,
                sheet_name,
                row,
                article is None,
                supply_name is None,
            )
        rows.append((article, supply_name))

    logger.info(
        "Article supplies Excel file read: path=%s rows_count=%s empty_articles=%s empty_supply_names=%s",
        path_xls,
        len(rows),
        empty_article_count,
        empty_supply_name_count,
    )
    return rows


def clear_art_supp():
    logger.info("Clearing article supplies")
    article_supply_repository.clear()
